import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# Excel dosyası oluştur
wb = Workbook()
ws = wb.active
ws.title = "Günlük Ajanda"

# Renk tanımlamaları
HEADER_COLOR = "1F4E78"  # Koyu mavi
SUBHEADER_COLOR = "2F75B5"  # Açık mavi
ALTERNATE_ROW_COLOR = "F2F7FB"  # Çok açık mavi
BORDER_COLOR = "B4C6E7"  # Pastel mavi

# Başlık oluştur
ws.merge_cells('A1:E1')
title_cell = ws['A1']
title_cell.value = f"Günlük İş Takip Ajandası - {datetime.now().strftime('%d.%m.%Y')}"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Alt başlıklar
headers = ["Saat", "Görev", "Öncelik", "Durum", "Notlar"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color=SUBHEADER_COLOR, end_color=SUBHEADER_COLOR, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Saat aralıkları oluştur (08:00 - 18:00)
start_time = datetime.strptime("08:00", "%H:%M")
time_slots = []
for i in range(21):  # 30'ar dakikalık aralıklarla 21 slot
    time_slots.append((start_time + timedelta(minutes=30*i)).strftime("%H:%M"))

# Verileri doldur
for row, time in enumerate(time_slots, 3):
    # Alternatif satır renklendirmesi
    row_color = ALTERNATE_ROW_COLOR if row % 2 == 0 else "FFFFFF"
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        if col == 1:
            cell.value = time
        elif col == 3:
            cell.value = "Normal"  # Varsayılan öncelik
        elif col == 4:
            cell.value = "Bekliyor"  # Varsayılan durum

# Sütun genişliklerini ayarla
column_widths = {
    'A': 12,  # Saat
    'B': 45,  # Görev
    'C': 15,  # Öncelik
    'D': 15,  # Durum
    'E': 35   # Notlar
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Kenarlık stilleri
thick_border = Border(
    left=Side(style='medium', color=BORDER_COLOR),
    right=Side(style='medium', color=BORDER_COLOR),
    top=Side(style='medium', color=BORDER_COLOR),
    bottom=Side(style='medium', color=BORDER_COLOR)
)

thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)

# Tüm hücrelere kenarlık ekle
for row in ws.iter_rows(min_row=1, max_row=len(time_slots)+2, min_col=1, max_col=5):
    for cell in row:
        if row[0].row == 1 or row[0].row == 2:  # Başlık ve alt başlıklar için kalın kenarlık
            cell.border = thick_border
        else:
            cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Öncelik ve Durum için veri doğrulama ve koşullu biçimlendirme
from openpyxl.worksheet.datavalidation import DataValidation

# Öncelik için veri doğrulama ve renklendirme
priority_dv = DataValidation(
    type="list",
    formula1='"Yüksek,Normal,Düşük"',
    allow_blank=True
)
ws.add_data_validation(priority_dv)
priority_dv.add(f'C3:C{len(time_slots)+2}')

# Durum için veri doğrulama ve renklendirme
status_dv = DataValidation(
    type="list",
    formula1='"Bekliyor,Devam Ediyor,Tamamlandı,İptal"',
    allow_blank=True
)
ws.add_data_validation(status_dv)
status_dv.add(f'D3:D{len(time_slots)+2}')

# Koşullu biçimlendirme kuralları
# Öncelik renklendirmesi
red_text = Font(color="FF0000")
green_text = Font(color="008000")
blue_text = Font(color="0000FF")

for row in range(3, len(time_slots) + 3):
    priority_cell = ws.cell(row=row, column=3)
    status_cell = ws.cell(row=row, column=4)
    
    # Öncelik renklendirmesi
    if priority_cell.value == "Yüksek":
        priority_cell.font = red_text
    elif priority_cell.value == "Normal":
        priority_cell.font = blue_text
    elif priority_cell.value == "Düşük":
        priority_cell.font = green_text
    
    # Durum renklendirmesi
    if status_cell.value == "Tamamlandı":
        status_cell.font = green_text
    elif status_cell.value == "Devam Ediyor":
        status_cell.font = blue_text
    elif status_cell.value == "İptal":
        status_cell.font = red_text

# Alt bilgi ekle
footer_row = len(time_slots) + 3
ws.merge_cells(f'A{footer_row}:E{footer_row}')
footer_cell = ws[f'A{footer_row}']
footer_cell.value = "© Günlük İş Takip Ajandası"
footer_cell.font = Font(italic=True, size=9, color="666666")
footer_cell.alignment = Alignment(horizontal="center")
footer_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# Sayfayı yazdırmaya hazırla
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToPage = True

# Excel dosyasını kaydet
wb.save('day.xlsx')
